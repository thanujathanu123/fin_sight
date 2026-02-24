import csv
import io
from datetime import datetime, timedelta
from typing import Dict, List, Any
from django.http import HttpResponse
from django.db.models import QuerySet, Count, Avg, Sum, Q
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Transaction, Alert, LedgerUpload, AuditLog, RiskProfile


class DataExporter:
    """Base class for data export functionality"""

    def __init__(self, queryset: QuerySet, fields: List[str] = None):
        self.queryset = queryset
        self.fields = fields or []

    def get_data(self) -> List[Dict[str, Any]]:
        """Extract data from queryset"""
        if self.fields:
            return list(self.queryset.values(*self.fields))
        return list(self.queryset.values())


class CSVExporter(DataExporter):
    """Export data to CSV format"""

    def export(self, filename: str = None) -> HttpResponse:
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'export_{timestamp}.csv'

        data = self.get_data()
        if not data:
            # Return empty CSV with headers
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            writer = csv.writer(response)
            if self.fields:
                writer.writerow(self.fields)
            return response

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(data[0].keys())  # Write headers

        for row in data:
            writer.writerow(row.values())

        return response


class ExcelExporter(DataExporter):
    """Export data to Excel format"""

    def export(self, filename: str = None, sheet_name: str = 'Data') -> HttpResponse:
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'export_{timestamp}.xlsx'

        data = self.get_data()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            # Add headers even for empty data
            if self.fields:
                for col, field in enumerate(self.fields, 1):
                    cell = ws.cell(row=1, column=col, value=field)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            else:
                ws.cell(row=1, column=1, value="No data available")
        else:
            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Write data
            for row, item in enumerate(data, 2):
                for col, (key, value) in enumerate(item.items(), 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    # Format risk scores
                    if key == 'risk_score' and isinstance(value, (int, float)):
                        if value >= 70:
                            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        elif value >= 40:
                            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class PDFExporter(DataExporter):
    """Export data to PDF format"""

    def __init__(self, queryset: QuerySet, fields: List[str] = None, title: str = "Data Export"):
        super().__init__(queryset, fields)
        self.title = title

    def export(self, filename: str = None) -> HttpResponse:
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'export_{timestamp}.pdf'

        data = self.get_data()

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )

        # Title
        elements.append(Paragraph(self.title, title_style))
        elements.append(Spacer(1, 12))

        if not data:
            elements.append(Paragraph("No data available for export.", styles['Normal']))
        else:
            # Prepare table data
            headers = list(data[0].keys())
            table_data = [headers]  # Header row

            for item in data:
                row = []
                for key in headers:
                    value = item[key]
                    # Format values
                    if isinstance(value, (int, float)) and key == 'risk_score':
                        row.append(f"{value:.1f}")
                    elif isinstance(value, datetime):
                        row.append(value.strftime('%Y-%m-%d %H:%M'))
                    else:
                        row.append(str(value) if value is not None else '')
                table_data.append(row)

            # Create table
            table = Table(table_data, repeatRows=1)

            # Table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])

            # Add risk score color coding
            risk_score_col = None
            for i, header in enumerate(headers):
                if header == 'risk_score':
                    risk_score_col = i
                    break

            if risk_score_col is not None:
                for row_idx, row in enumerate(table_data[1:], 1):  # Skip header
                    try:
                        risk_value = float(row[risk_score_col])
                        if risk_value >= 70:
                            table_style.add('BACKGROUND', (risk_score_col, row_idx), (risk_score_col, row_idx), colors.red)
                        elif risk_value >= 40:
                            table_style.add('BACKGROUND', (risk_score_col, row_idx), (risk_score_col, row_idx), colors.yellow)
                    except (ValueError, IndexError):
                        pass

            table.setStyle(table_style)
            elements.append(table)

        # Build PDF
        doc.build(elements)

        # Create response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


# Specialized exporters for different data types

class TransactionExporter:
    """Export transactions with specialized formatting"""

    @staticmethod
    def export_csv(queryset: QuerySet, filename: str = None) -> HttpResponse:
        fields = ['date', 'amount', 'description', 'category', 'risk_score', 'status']
        exporter = CSVExporter(queryset, fields)
        return exporter.export(filename or 'transactions.csv')

    @staticmethod
    def export_excel(queryset: QuerySet, filename: str = None) -> HttpResponse:
        fields = ['date', 'amount', 'description', 'category', 'risk_score', 'status', 'reference_id']
        exporter = ExcelExporter(queryset, fields)
        return exporter.export(filename or 'transactions.xlsx', 'Transactions')

    @staticmethod
    def export_pdf(queryset: QuerySet, filename: str = None) -> HttpResponse:
        fields = ['date', 'amount', 'description', 'category', 'risk_score', 'status']
        exporter = PDFExporter(queryset, fields, 'Transaction Report')
        return exporter.export(filename or 'transactions.pdf')


class AlertExporter:
    """Export alerts with specialized formatting"""

    @staticmethod
    def export_csv(queryset: QuerySet, filename: str = None) -> HttpResponse:
        fields = ['title', 'description', 'severity', 'status', 'created_at', 'transaction__reference_id']
        exporter = CSVExporter(queryset, fields)
        return exporter.export(filename or 'alerts.csv')

    @staticmethod
    def export_excel(queryset: QuerySet, filename: str = None) -> HttpResponse:
        fields = ['title', 'description', 'severity', 'status', 'created_at', 'transaction__reference_id', 'assigned_to__username']
        exporter = ExcelExporter(queryset, fields)
        return exporter.export(filename or 'alerts.xlsx', 'Alerts')

    @staticmethod
    def export_pdf(queryset: QuerySet, filename: str = None) -> HttpResponse:
        fields = ['title', 'description', 'severity', 'status', 'created_at']
        exporter = PDFExporter(queryset, fields, 'Alert Report')
        return exporter.export(filename or 'alerts.pdf')


class AnalyticsReportExporter:
    """Export analytics reports"""

    @staticmethod
    def generate_summary_report(start_date: datetime = None, end_date: datetime = None) -> Dict[str, Any]:
        """Generate summary analytics data"""
        if not start_date:
            start_date = timezone.now() - timedelta(days=30)
        if not end_date:
            end_date = timezone.now()

        # Transaction analytics
        transactions = Transaction.objects.filter(date__gte=start_date, date__lte=end_date)
        transaction_stats = transactions.aggregate(
            total_count=Count('id'),
            total_amount=Sum('amount'),
            avg_risk=Avg('risk_score'),
            high_risk_count=Count('id', filter=Q(risk_score__gte=70))
        )

        # Alert analytics
        alerts = Alert.objects.filter(created_at__gte=start_date, created_at__lte=end_date)
        alert_stats = alerts.aggregate(
            total_alerts=Count('id'),
            resolved_alerts=Count('id', filter=Q(status='resolved')),
            critical_alerts=Count('id', filter=Q(severity='critical'))
        )

        # Risk distribution
        risk_distribution = {
            'low': transactions.filter(risk_score__lt=40).count(),
            'medium': transactions.filter(risk_score__gte=40, risk_score__lt=70).count(),
            'high': transactions.filter(risk_score__gte=70).count()
        }

        return {
            'period': {'start': start_date, 'end': end_date},
            'transactions': transaction_stats,
            'alerts': alert_stats,
            'risk_distribution': risk_distribution,
            'generated_at': timezone.now()
        }

    @staticmethod
    def export_summary_pdf(start_date: datetime = None, end_date: datetime = None, filename: str = None) -> HttpResponse:
        """Export summary report as PDF"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'analytics_summary_{timestamp}.pdf'

        data = AnalyticsReportExporter.generate_summary_report(start_date, end_date)

        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=30, alignment=1)
        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=14, spaceAfter=20)
        normal_style = styles['Normal']

        # Title
        elements.append(Paragraph("FinSight Analytics Summary Report", title_style))
        elements.append(Spacer(1, 12))

        # Period
        period_text = f"Report Period: {data['period']['start'].strftime('%Y-%m-%d')} to {data['period']['end'].strftime('%Y-%m-%d')}"
        elements.append(Paragraph(period_text, normal_style))
        elements.append(Paragraph(f"Generated: {data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        elements.append(Spacer(1, 20))

        # Transaction Summary
        elements.append(Paragraph("Transaction Summary", heading_style))
        trans_data = [
            ['Metric', 'Value'],
            ['Total Transactions', str(data['transactions']['total_count'] or 0)],
            ['Total Amount', f"${data['transactions']['total_amount'] or 0:,.2f}"],
            ['Average Risk Score', f"{data['transactions']['avg_risk'] or 0:.1f}"],
            ['High Risk Transactions', str(data['transactions']['high_risk_count'] or 0)]
        ]
        trans_table = Table(trans_data)
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(trans_table)
        elements.append(Spacer(1, 20))

        # Alert Summary
        elements.append(Paragraph("Alert Summary", heading_style))
        alert_data = [
            ['Metric', 'Value'],
            ['Total Alerts', str(data['alerts']['total_alerts'] or 0)],
            ['Resolved Alerts', str(data['alerts']['resolved_alerts'] or 0)],
            ['Critical Alerts', str(data['alerts']['critical_alerts'] or 0)]
        ]
        alert_table = Table(alert_data)
        alert_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(alert_table)
        elements.append(Spacer(1, 20))

        # Risk Distribution
        elements.append(Paragraph("Risk Distribution", heading_style))
        risk_data = [
            ['Risk Level', 'Count'],
            ['Low Risk (0-39)', str(data['risk_distribution']['low'])],
            ['Medium Risk (40-69)', str(data['risk_distribution']['medium'])],
            ['High Risk (70-100)', str(data['risk_distribution']['high'])]
        ]
        risk_table = Table(risk_data)
        risk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(risk_table)

        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response